import requests
import openpyxl
from bs4 import BeautifulSoup as BS
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter


def main():
    wb = Workbook()
    worksheet = wb.active
    wb.save('notebooks.xlsx')
    side = Side(border_style='thin')
    border = Border(left=side,
                    right=side,
                    top=side,
                    bottom=side)
    aligment = Alignment(horizontal='center',
                         vertical='center')
    column_widths = []

    headers = {
        "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9",
        "user-agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.106 Safari/537.36"
    }

    colomn_names = ['Название',
                    'Цена',
                    'Наличие',
                    'Процессор',
                    'Видеокарта',
                    'Оперативная память',
                    'Жесткий диск',
                    'SSD накопитель',
                    'Дисплей',
                    'Операционная система']
    for colomn, name in enumerate(colomn_names, 1):
        cell = worksheet.cell(column=colomn,
                              row=1,
                              value=name)
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = aligment
        column_widths.append(len(name))

    page = 1
    row = 2
    while True:
        r = requests.get(f'https://delta-game.ru/product-category/notebooks/page/{page}', headers=headers)
        print(f'https://delta-game.ru/product-category/notebooks/page/{page}')
        html = BS(r.content, 'html.parser')
        notebooks = html.select('.row > .products > .col-md-3')
        if len(notebooks):
            for notebook_raw in notebooks:
                notebook = dict()
                notebook['Название'] = notebook_raw.select('a > h2')[0].text
                notebook['Цена'] = int(notebook_raw.select('bdi')[0].text[:-2].replace(' ', ''))
                notebook['Наличие'] = notebook_raw.select('.stock-status')[0].text

                component_names = [name.text.capitalize() for name in notebook_raw.select('.pc_part')]
                component_models = [component.text for component in notebook_raw.select('.pc_part_model')]
                for model, name in zip(component_models, component_names):
                    notebook[name] = model
                if 'Ssd накопитель' in notebook.keys():
                    notebook['SSD накопитель'] = notebook.pop('Ssd накопитель')
                for column, name in enumerate(colomn_names, 1):
                    cell = worksheet.cell(column=column,
                                          row=row,
                                          value=notebook.get(name, 'Нет'))
                    cell.border = border
                    cell.alignment = aligment
                    column_widths[column - 1] = max(column_widths[column - 1], len(str(cell.value)))
                row += 1
            wb.save('notebooks.xlsx')
            page += 1
        else:
            break
    for i, column_width in enumerate(column_widths, 1):
        worksheet.column_dimensions[get_column_letter(i)].width = column_width + 1
    wb.save('notebooks.xlsx')


if __name__ == '__main__':
    main()
